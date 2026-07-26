#!/usr/bin/env python3
"""
READ-ONLY: propose a WooCommerce-product ↔ QuickBooks-item mapping so the connector
can match by SKU later. Matches WC product NAME against QBO item DESCRIPTION
(both bilingual), weighting the Chinese name highest. Outputs an Excel for
Michelle's team to verify/correct (esp. loose/by-piece + ambiguous grade-variants).

NO writes to QuickBooks. Output only.
"""
import json, re, difflib, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

PROJ = os.path.dirname(os.path.abspath(__file__))
QDIR = os.path.join(PROJ, "qbo-backups/2026-06-04T08-03-52Z")
WDIR = os.path.join(PROJ, "wc-backups/2026-06-04T08-04-29Z/full-store")

items = json.load(open(f"{QDIR}/Item.json"))
prods = json.load(open(f"{WDIR}/products.json"))
# active QBO items with a description, and not the obvious "- copy" dupes deprioritised
qbo = [i for i in items if i.get("Active") and (i.get("Description") or i.get("Name"))]

CJK = lambda s: "".join(c for c in (s or "") if "一" <= c <= "鿿")
def eng_tokens(s):
    return set(t for t in re.split(r"[^A-Z0-9]+", (s or "").upper()) if len(t) > 1)

def desc(i): return i.get("Description") or i.get("Name") or ""

# pre-index QBO
qidx = [{"item": i, "text": desc(i), "cn": CJK(desc(i)), "eng": eng_tokens(desc(i))} for i in qbo]

def score(wc_name, q):
    cn_w = CJK(wc_name); cn_q = q["cn"]
    if cn_w and cn_q:
        if cn_w == cn_q: cn = 1.0
        elif cn_w in cn_q or cn_q in cn_w: cn = 0.85
        else: cn = difflib.SequenceMatcher(None, cn_w, cn_q).ratio()
    else:
        cn = 0.0
    ew, eq = eng_tokens(wc_name), q["eng"]
    eng = len(ew & eq) / len(ew | eq) if (ew | eq) else 0.0
    has_cn = bool(cn_w and cn_q)
    return round((0.6*cn + 0.4*eng) if has_cn else (0.3 + 0.7*eng), 3)

rows = []
for p in prods:
    name = p.get("name", ""); sku = p.get("sku", "")
    scored = sorted(((score(name, q), q) for q in qidx), key=lambda x: -x[0])
    top_s, top_q = scored[0]
    second_s = scored[1][0] if len(scored) > 1 else 0
    ambiguous = (top_s - second_s) < 0.05 and top_s > 0.4
    conf = ("HIGH" if top_s >= 0.85 else "MEDIUM" if top_s >= 0.6 else "LOW" if top_s >= 0.4 else "NONE")
    alts = "; ".join(f"{q['item'].get('Name')}:{q['text'][:30]}" for s, q in scored[1:4] if s > 0.4)
    rows.append({
        "wc_sku": sku, "wc_name": name,
        "qbo_code": top_q["item"].get("Name"), "qbo_desc": top_q["text"],
        "score": top_s, "conf": conf, "ambiguous": "YES" if ambiguous else "",
        "alts": alts,
    })

from collections import Counter
c = Counter(r["conf"] for r in rows)
amb = sum(1 for r in rows if r["ambiguous"])
print(f"WC products: {len(rows)}")
print(f"  HIGH (auto-OK): {c['HIGH']}  MEDIUM (review): {c['MEDIUM']}  LOW (check): {c['LOW']}  NONE (manual): {c['NONE']}")
print(f"  ambiguous (multiple close QBO matches): {amb}")

# Excel
wb = Workbook(); ws = wb.active; ws.title = "Product Mapping"
hdr = ["WC SKU", "Website product name", "→", "Proposed QBO code", "QBO description", "Match", "Confidence", "Ambiguous?", "Other possible QBO matches", "✔ Correct? (Y/N)", "If wrong, correct QBO code"]
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="2E7D32")
for i, h in enumerate(hdr, 1):
    cell = ws.cell(1, i, h); cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(wrap_text=True, vertical="top")
conf_fill = {"HIGH": PatternFill("solid", fgColor="C8E6C9"), "MEDIUM": PatternFill("solid", fgColor="FFF9C4"),
             "LOW": PatternFill("solid", fgColor="FFE0B2"), "NONE": PatternFill("solid", fgColor="FFCDD2")}
yellow = PatternFill("solid", fgColor="FFF59D")
for r in rows:
    ws.append([r["wc_sku"], r["wc_name"], "→", r["qbo_code"], r["qbo_desc"], r["score"], r["conf"], r["ambiguous"], r["alts"], "", ""])
    row = ws.max_row
    ws.cell(row, 7).fill = conf_fill.get(r["conf"])
    ws.cell(row, 10).fill = yellow; ws.cell(row, 11).fill = yellow
for col, w in zip("ABCDEFGHIJK", [20, 42, 3, 14, 38, 7, 11, 10, 40, 14, 18]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"
out = os.path.join(PROJ, "QBO-WC-Product-Mapping-DRAFT-2026-06-04.xlsx")
wb.save(out)
print("Excel:", out)
